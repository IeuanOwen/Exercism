from openpyxl import load_workbook
from bs4 import BeautifulSoup
import requests
import re

pattern = r'mgid'

workbook = load_workbook('C:\\Users\IFPI\Desktop\listexport.xlsx')
sheet = workbook.active
row_count = sheet.max_row
column = sheet.max_column
for i in range(row_count):
   try:
      site = (sheet.cell(row=i+1, column=column).value)
      r = requests.get("https//" + site)
      soup = r.text
      if re.search(pattern, soup) == True:
         print('yes')
      else:
         print('no')
   except Exception:
      continue



